import os
import pandas as pd
from binance.client import Client
import openpyxl
from datetime import datetime, timedelta
import keys  # Ensure you have your API keys in a file named keys.py

# Your Binance API credentials
api_key = keys.api
api_secret = keys.secret

# Initialize the Binance client
client = Client(api_key, api_secret)

# Path to the past data CSV file
past_data_file = 'past_data.csv'

# Fetch the margin balance information from Binance Futures
def get_margin_balance():
    account_info = client.futures_account()
    margin_balance = float(account_info['totalMarginBalance'])
    return margin_balance

# Initialize Excel file with past data if it doesn't exist
def initialize_excel(file_path, past_data_file):
    if not os.path.exists(file_path):
        past_data = pd.read_csv(past_data_file)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Records"
        ws.append(['Execution Date Time', 'Margin Balance', 'PNL', 'PNL (%)', 'Total PNL', 'Total PNL (%)'])
        for index, row in past_data.iterrows():
            ws.append([row['Execution Date Time'], row['Margin Balance'], row['PNL'], row['PNL (%)'], row['Total PNL'], row['Total PNL (%)']])
        wb.save(file_path)

# Update Excel file with the new balance information
def update_excel(file_path):
    if not os.path.exists(file_path):
        initialize_excel(file_path, past_data_file)

    wb = openpyxl.load_workbook(file_path)
    record_ws = wb['Records']

    # Fetch the current margin balance
    margin_balance = get_margin_balance()
    execution_date_time = datetime.now().strftime("%b %d %H:%M")

    # Calculate PNL and PNL (%)
    last_row = record_ws.max_row
    if last_row > 1:
        last_balance = record_ws.cell(last_row, 2).value
        pnl = margin_balance - last_balance
        pnl_percentage = (pnl / last_balance) * 100 if last_balance != 0 else 0
    else:
        pnl = 0  # Initial entry PNL is 0
        pnl_percentage = 0

    # Calculate total PNL and total PNL (%)
    if last_row > 1:
        first_balance = record_ws.cell(2, 2).value
        total_pnl = margin_balance - first_balance
        total_pnl_percentage = (total_pnl / first_balance) * 100 if first_balance != 0 else 0
    else:
        total_pnl = pnl
        total_pnl_percentage = pnl_percentage

    record_ws.append([execution_date_time, margin_balance, pnl, pnl_percentage, total_pnl, total_pnl_percentage])
    wb.save(file_path)

    # Calculate average PNLs
    calculate_average_pnl(file_path)

# Calculate average PNL for 8 hours and 1 day intervals
def calculate_average_pnl(file_path):
    df = pd.read_excel(file_path, sheet_name='Records')
    df['Execution Date Time'] = pd.to_datetime(df['Execution Date Time'], format="%b %d %H:%M")

    now = datetime.now()

    # Calculate time intervals
    first_time = df['Execution Date Time'].min()
    time_elapsed = (now - first_time).total_seconds() / 3600  # Time elapsed in hours

    if time_elapsed == 0:
        avg_pnl_8_hours = avg_pnl_1_day = 0
        avg_pnl_8_hours_percentage = avg_pnl_1_day_percentage = 0
    else:
        # Calculate average PNL
        total_pnl = df['Total PNL'].iloc[-1]
        avg_pnl_8_hours = (total_pnl * 8) / time_elapsed
        avg_pnl_1_day = (total_pnl * 24) / time_elapsed

        # Calculate average PNL percentage
        first_balance = df['Margin Balance'].iloc[0]
        avg_pnl_8_hours_percentage = (avg_pnl_8_hours / first_balance) * 100 if first_balance != 0 else 0
        avg_pnl_1_day_percentage = (avg_pnl_1_day / first_balance) * 100 if first_balance != 0 else 0

    # Output the results
    print(f"Average PNL for the last 8 hours: {avg_pnl_8_hours} ({avg_pnl_8_hours_percentage}%)")
    print(f"Average PNL for the last 1 day: {avg_pnl_1_day} ({avg_pnl_1_day_percentage}%)")

    # Update Excel with average PNLs
    wb = openpyxl.load_workbook(file_path)
    if 'Averages' not in wb.sheetnames:
        avg_ws = wb.create_sheet(title="Averages")
        avg_ws.append(['Interval', 'Average PNL', 'Average PNL (%)'])
    else:
        avg_ws = wb['Averages']
    avg_ws.append(['Last 8 Hours', avg_pnl_8_hours, avg_pnl_8_hours_percentage])
    avg_ws.append(['Last 1 Day', avg_pnl_1_day, avg_pnl_1_day_percentage])
    wb.save(file_path)

# Main function to fetch and update positions
def main():
    if os.name == 'nt':
        desktop_path = os.path.join(os.environ['USERPROFILE'], 'Desktop', 'margin_pnl.xlsx')
    else:
        desktop_path = os.path.join(os.environ['HOME'], 'Desktop', 'margin_pnl.xlsx')

    update_excel(desktop_path)

if __name__ == "__main__":
    main()
